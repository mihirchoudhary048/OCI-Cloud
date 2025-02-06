import oci
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference

# Load OCI configuration
config = oci.config.from_file("~/.oci/config")

# Initialize OCI clients
identity_client = oci.identity.IdentityClient(config)
virtual_network_client = oci.core.VirtualNetworkClient(config)
compute_client = oci.core.ComputeClient(config)
block_storage_client = oci.core.BlockstorageClient(config)
object_storage_client = oci.object_storage.ObjectStorageClient(config)
database_client = oci.database.DatabaseClient(config)
load_balancer_client = oci.load_balancer.LoadBalancerClient(config)
cloud_advisor_client = oci.optimizer.OptimizerClient(config)
cloud_guard_client = oci.cloud_guard.CloudGuardClient(config)
namespace = object_storage_client.get_namespace().data

# Get tenancy ID
tenancy_id = config["tenancy"]

# Initialize result storage
resources = {}
findings = {}
cloud_advisor_recommendations = []
cloud_guard_findings = []

try:
    # Fetch all compartments
    compartments = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    compartments.append(oci.identity.models.Compartment(id=tenancy_id, name="Tenancy Root"))

    # Discover resources in each compartment
    for compartment in compartments:
        if compartment.lifecycle_state == "ACTIVE":
            print(f"Discovering resources in compartment: {compartment.name}")
            resources[compartment.name] = {}
            findings[compartment.name] = []

            # Discover VCNs
            vcn_response = oci.pagination.list_call_get_all_results(
                virtual_network_client.list_vcns,
                compartment_id=compartment.id
            ).data
            vcn_findings = []
            for vcn in vcn_response:
                resources[compartment.name].setdefault("VCNs", []).append({"name": vcn.display_name, "id": vcn.id})
                # Best practice: Check for wide CIDR ranges
                if vcn.cidr_block == "0.0.0.0/0":
                    vcn_findings.append(f"VCN '{vcn.display_name}' has an open CIDR block.")
            findings[compartment.name].extend(vcn_findings)

            # Discover Compute Instances
            instance_response = oci.pagination.list_call_get_all_results(
                compute_client.list_instances,
                compartment_id=compartment.id
            ).data
            instance_findings = []
            for instance in instance_response:
                resources[compartment.name].setdefault("Compute Instances", []).append({
                    "name": instance.display_name,
                    "id": instance.id
                })

                # Check if instance is using the latest platform images
                image_details = compute_client.get_image(instance.image_id).data
                if "platform" in image_details.operating_system and not image_details.is_latest:
                    instance_findings.append(f"Instance '{instance.display_name}' is not using the latest platform image.")

                # Check for SSH key-based authentication
                if not instance.metadata or "ssh_authorized_keys" not in instance.metadata:
                    instance_findings.append(f"Instance '{instance.display_name}' does not have SSH key-based authentication configured.")

                # Check if password-based login is disabled
                if instance.metadata and "disable_password_auth" not in instance.metadata:
                    instance_findings.append(f"Instance '{instance.display_name}' has password-based login enabled.")

                # Check for logging agents
                if "logging_agent" not in instance.metadata or instance.metadata.get("logging_agent") != "configured":
                    instance_findings.append(f"Instance '{instance.display_name}' does not have logging agents configured.")

                # Check if NSGs restrict unnecessary ports
                vnics = compute_client.list_vnic_attachments(compartment_id=compartment.id, instance_id=instance.id).data
                for vnic_attachment in vnics:
                    vnic = virtual_network_client.get_vnic(vnic_attachment.vnic_id).data
                    nsgs = vnic.nsg_ids
                    for nsg_id in nsgs:
                        try:
                            nsg_rules = oci.pagination.list_call_get_all_results(
                                virtual_network_client.list_network_security_group_security_rules,
                                network_security_group_id=nsg_id
                            ).data
                            for rule in nsg_rules:
                                if rule.direction == "INGRESS" and rule.source == "0.0.0.0/0":
                                    instance_findings.append(f"Instance '{instance.display_name}' NSG allows unrestricted ingress.")
                        except oci.exceptions.ServiceError as e:
                            instance_findings.append(f"Error fetching rules for NSG ID {nsg_id}: {str(e)}")

            findings[compartment.name].extend(instance_findings)

            # Discover Block Volumes
            volume_response = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volumes,
                compartment_id=compartment.id
            ).data
            volume_findings = []
            for volume in volume_response:
                resources[compartment.name].setdefault("Block Volumes", []).append({
                    "name": volume.display_name,
                    "id": volume.id
                })
                # Check if the volume is attached to any instance
                attachments = oci.pagination.list_call_get_all_results(
                    compute_client.list_volume_attachments,
                    compartment_id=compartment.id,
                    volume_id=volume.id
                ).data
                if not attachments:  # No attachments found
                    volume_findings.append(f"Volume '{volume.display_name}' is not attached to any instance.")
                # Best practice: Ensure backup policy is set
                if not volume.is_auto_tune_enabled:
                    volume_findings.append(f"Volume '{volume.display_name}' does not have auto-tune enabled.")
            findings[compartment.name].extend(volume_findings)

            # Discover Object Storage Buckets
            bucket_response = oci.pagination.list_call_get_all_results(
                object_storage_client.list_buckets,
                namespace_name=namespace,
                compartment_id=compartment.id
            ).data
            bucket_findings = []
            for bucket in bucket_response:
                resources[compartment.name].setdefault("Buckets", []).append({"name": bucket.name})
                # Fetch detailed bucket info to check for public access
                bucket_details = object_storage_client.get_bucket(
                    namespace_name=namespace,
                    bucket_name=bucket.name
                ).data
                # Best practice: Check for public access
                if bucket_details.public_access_type != "NoPublicAccess":
                    bucket_findings.append(f"Bucket '{bucket.name}' allows public access.")
                # Discover Objects in Buckets
                object_response = oci.pagination.list_call_get_all_results(
                    object_storage_client.list_objects,
                    namespace_name=namespace,
                    bucket_name=bucket.name
                ).data
                resources[compartment.name].setdefault("Bucket Objects", []).extend([
                    {"bucket_name": bucket.name, "object_name": obj.name} for obj in object_response.objects
                ])
            findings[compartment.name].extend(bucket_findings)

            # Discover Autonomous Databases
            adb_response = oci.pagination.list_call_get_all_results(
                database_client.list_autonomous_databases,
                compartment_id=compartment.id
            ).data
            adb_findings = []
            for adb in adb_response:
                resources[compartment.name].setdefault("Autonomous Databases", []).append({
                    "name": adb.display_name,
                    "id": adb.id
                })
                # Best practice: Check for appropriate workload type
                if adb.db_workload != "OLTP":
                    adb_findings.append(f"ADB '{adb.display_name}' is not optimized for OLTP workloads.")
            findings[compartment.name].extend(adb_findings)

            # Discover Load Balancers
            lb_response = oci.pagination.list_call_get_all_results(
                load_balancer_client.list_load_balancers,
                compartment_id=compartment.id
            ).data
            lb_findings = []
            for lb in lb_response:
                resources[compartment.name].setdefault("Load Balancers", []).append({
                    "name": lb.display_name,
                    "id": lb.id
                })
                # Best practice: Ensure SSL termination is configured
                if not lb.shape_name.startswith("flexible"):
                    lb_findings.append(f"Load Balancer '{lb.display_name}' is not using a flexible shape.")
            findings[compartment.name].extend(lb_findings)

    # Discover Cloud Advisor Recommendations
    try:
        advisor_recommendations = oci.pagination.list_call_get_all_results(
            cloud_advisor_client.list_recommendations,
            compartment_id=tenancy_id,
            compartment_id_in_subtree=True  # Include sub-compartments
        ).data
        for recommendation in advisor_recommendations:
            cloud_advisor_recommendations.append({
                "Name": recommendation.name,
                "Recommendation": getattr(recommendation, "description", "No description available")
            })
    except oci.exceptions.ServiceError as e:
        print(f"Cloud Advisor Service Error: {e}")

    # Discover Cloud Guard Findings
    try:
        cloud_guard_problems = oci.pagination.list_call_get_all_results(
            cloud_guard_client.list_problems,
            compartment_id=tenancy_id,
            compartment_id_in_subtree=True
        ).data
        for problem in cloud_guard_problems:
            cloud_guard_findings.append({
                "Name": problem.resource_name,
                "Description": problem.labels
            })
    except oci.exceptions.ServiceError as e:
        print(f"Cloud Guard Service Error: {e}")

    # Export data to JSON
    with open("oci_resources.json", "w") as file:
        json.dump({"resources": resources, "findings": findings, "cloud_advisor_recommendations": cloud_advisor_recommendations, "cloud_guard_findings": cloud_guard_findings}, file, indent=4)

    print("Resource discovery and validation completed. Results saved to 'oci_resources.json'.")

    # Export data to Excel
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Findings Summary"

    # Add findings summary
    summary_sheet.append(["Compartment", "Issue", "Recommendation"])
    for compartment, issues in findings.items():
        for issue in issues:
            summary_sheet.append([compartment, issue, "Refer to OCI best practices."])

    # Style misconfigurations
    for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            cell.font = Font(bold=True)

    # Add Cloud Advisor Recommendations
    advisor_sheet = workbook.create_sheet(title="Cloud Advisor")
    advisor_sheet.append(["Name", "Recommendation"])
    if cloud_advisor_recommendations:
        for recommendation in cloud_advisor_recommendations:
            advisor_sheet.append([recommendation["Name"], recommendation["Recommendation"]])
    else:
        advisor_sheet.append(["No Cloud Advisor recommendations found."])

    # Add Cloud Guard Findings
    cloud_guard_sheet = workbook.create_sheet(title="Cloud Guard")
    cloud_guard_sheet.append(["Resource Name", "Description"])
    if cloud_guard_findings:
        for finding in cloud_guard_findings:
            cloud_guard_sheet.append([finding["Name"], finding["Description"]])
    else:
        cloud_guard_sheet.append(["No Cloud Guard findings found."])

    # Add data sheets for each resource type
    for resource_type in ["VCNs", "Compute Instances", "Block Volumes", "Buckets", "Bucket Objects", "Autonomous Databases", "Load Balancers"]:
        sheet = workbook.create_sheet(title=resource_type)
        sheet.append(["Compartment", "Name", "ID"])
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):
                sheet.append([compartment, item.get("name"), item.get("id", "N/A")])

    # Add Visualization Sheet
    visualization_sheet = workbook.create_sheet(title="Visualizations")
    visualization_sheet.append(["Resource Type", "Count"])

    resource_counts = {
        "VCNs": sum(len(data.get("VCNs", [])) for data in resources.values()),
        "Compute Instances": sum(len(data.get("Compute Instances", [])) for data in resources.values()),
        "Block Volumes": sum(len(data.get("Block Volumes", [])) for data in resources.values()),
        "Buckets": sum(len(data.get("Buckets", [])) for data in resources.values()),
        "Bucket Objects": sum(len(data.get("Bucket Objects", [])) for data in resources.values()),
        "Autonomous Databases": sum(len(data.get("Autonomous Databases", [])) for data in resources.values()),
        "Load Balancers": sum(len(data.get("Load Balancers", [])) for data in resources.values()),
    }

    for resource_type, count in resource_counts.items():
        visualization_sheet.append([resource_type, count])

    pie_chart = PieChart()
    pie_chart.title = "Resource Distribution"
    pie_data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(resource_counts) + 1)
    pie_labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(resource_counts) + 1)
    pie_chart.add_data(pie_data, titles_from_data=False)
    pie_chart.set_categories(pie_labels)
    visualization_sheet.add_chart(pie_chart, "D2")

    bar_chart = BarChart()
    bar_chart.title = "Resource Counts"
    bar_data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(resource_counts) + 1)
    bar_labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(resource_counts) + 1)
    bar_chart.add_data(bar_data, titles_from_data=False)
    bar_chart.set_categories(bar_labels)
    visualization_sheet.add_chart(bar_chart, "D20")

    # Save the Excel workbook
    workbook.save("oci_resources.xlsx")
    print("Detailed findings and visualizations saved to 'oci_resources.xlsx'.")

except oci.exceptions.ServiceError as e:
    print(f"Service Error: {e}")
except Exception as e:
    print(f"Unexpected Error: {e}")

